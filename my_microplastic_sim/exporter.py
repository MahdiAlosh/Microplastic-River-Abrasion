import pandas as pd
from openpyxl import load_workbook

def save_results_to_excel(input_file, output_file, all_results):
    data = []

    for water_type, polymers_data in all_results.items():
        for polymer, values in polymers_data.items():
            data.append({
                'Water Type': water_type,
                'Polymer': polymer,
                'Final Diameter Min (µm)': values['final_diameter_min'],
                'Final Diameter Max (µm)': values['final_diameter_max'],
                'Total Diameter Loss Min (µm)': values['total_loss_diameter_min'],
                'Total Diameter Loss Max (µm)': values['total_loss_diameter_max'],
                'Total Mass Loss Min (mg)': values['total_mass_loss_min'],
                'Total Mass Loss Max (mg)': values['total_mass_loss_max'],
                'Total Volume Loss Min (µm³)': values['total_volume_loss_min'],
                'Total Volume Loss Max (µm³)': values['total_volume_loss_max'],
                'Remaining Km Min': values['remaining_km_min'],
                'Remaining Km Max': values['remaining_km_max']
            })
    try:
        df = pd.DataFrame(data)
        if input_file == output_file:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                book = load_workbook(output_file)
                if 'Polymer_Results' in book.sheetnames:
                    del book['Polymer_Results']
                book.save(output_file)
                df.to_excel(writer, sheet_name='Polymer_Results', index=False)
        else:
            df.to_excel(output_file, sheet_name='Polymer_Results', index=False)

    except Exception as e:
        print(f"Error saving results to Excel: {e}")

