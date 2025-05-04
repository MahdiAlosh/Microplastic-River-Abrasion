import math
import os
from openpyxl import load_workbook
import pandas as pd

# Polymer properties (density in g/cm³)
polymers = {
    "ps_tio": {"a": 1.00036, "density": 1.05},
    "ps": {"a": 0.81021, "density": 1.05},
    "petg": {"a": 0.52683, "density": 1.27},
    "pet": {"a": 0.35373, "density": 1.39},
    "pa6": {"a": 0.32447, "density": 1.14}
}

def load_excel_data(file_path):
    """Load data from Excel file and return as DataFrame"""
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} not found!")
        return None
    try:
        # Read Excel file
        df = pd.read_excel(file_path) # DataFrame
        df = df.iloc[:-1]
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def simulate_abrasion(total_length, w_eff_min, w_eff_max):
    
    results = {}
    
    for polymer in polymers:
        # Simulation parameters
        initial_diameter_um = 5000  # 5 mm in µm
        min_diameter_um = max_diameter_um = 30  # Abruchkriterium
        river_length_km = total_length
        # w_eff_min = 14.8226336925046    # w/m²
        # w_eff_max = 1091.790083         # w/m²
        
        density = polymers[polymer]["density"]
        a = polymers[polymer]["a"]
        
        diameter_um_min = initial_diameter_um
        remaining_km_min = river_length_km
        step_min = 0
        sum_volume_loss_um3_min = sum_volume_loss_um3_max = sum_mass_loss_mg_min = sum_mass_loss_mg_max = 0

        # print(f"*********** Simulation für {polymer} ***********")
        # print(f"Startdurchmesser: {diameter_um_min} µm")
        initial_volume_um3_min = (math.pi/6) * diameter_um_min**3
        # print(f"Initial Volume (min): {initial_volume_um3_min:.2f} µm³")
        # print("="*50)

        while (diameter_um_min >= min_diameter_um) and (remaining_km_min > 0):
            step_min += 1
            # Convert to meters
            d_m = diameter_um_min * 1e-6
            
            # 1. Calculate contact area (m²)
            A = 0.5 * math.pi * d_m**2
            
            # 2. Calculate mass loss (mg/km)
            mass_loss_mg_min = (a * w_eff_min * 1000 * A)
            
            # 3. Calculate volume loss (µm³/km)
            volume_loss_um3_min = (mass_loss_mg_min * 1e9) / density  # mg → µg → µm³
            
            # 4. Update volume (µm³)
            initial_volume_um3_min = (math.pi/6) * diameter_um_min**3
            # print(f"Initial Volume (min): {initial_volume_um3_min:.2f} µm³")
            new_volume_um3_min = initial_volume_um3_min - volume_loss_um3_min
            # print(f"New Volume (min): {new_volume_um3_min:.2f} µm³")
            
            if new_volume_um3_min <= 0:
                # print(f"Partikel vollständig abradiert nach {step_min} Schritten")
                break
                
            # 5. Calculate new diameter (µm)
            new_diameter_um_min = (6 * new_volume_um3_min / math.pi) ** (1/3)
            
            # Progress output
            """ 
            print(f"Schritt {step_min}:")
            print(f"• Durchmesser (min): {diameter_um_min:.2f} → {new_diameter_um_min:.2f} µm")
            print(f"• Massenverlust (min): {mass_loss_mg_min:.6f} mg/km")
            print(f"• Volumenverlust (min): {volume_loss_um3_min:.2f} µm³")
            print("-"*50)
            """
            # sum up volume loss (min)
            sum_volume_loss_um3_min += volume_loss_um3_min

            # sum up mass loss (min)
            sum_mass_loss_mg_min += mass_loss_mg_min

            # Update for next iteration
            diameter_um_min = new_diameter_um_min # um -> µm

            remaining_km_min -= 1


        # Now calculate the maximum diameter
        # print("*"*50)
        diameter_um_max = initial_diameter_um
        remaining_km_max = river_length_km
        step_max = 0
        # print(f"\nSimulation für {polymer} über {river_length_km} km")
        # print(f"Startdurchmesser: {diameter_um_max} µm")
        initial_volume_um3_max = (math.pi/6) * diameter_um_max**3
        # print(f"Initial Volume (max): {initial_volume_um3_max:.2f} µm³")
        # print("="*50)

        while (diameter_um_max >= max_diameter_um) and (remaining_km_max > 0):
            step_max += 1
            # Convert to meters
            d_m = diameter_um_max * 1e-6
            
            # 1. Calculate contact area (m²)
            A = 0.5 * math.pi * d_m**2
            
            # 2. Calculate mass loss (mg/km)
            mass_loss_mg_max = (a * w_eff_max * 1000 * A)
            
            # 3. Calculate volume loss (µm³/km)
            volume_loss_um3_max = (mass_loss_mg_max * 1e9) / density
            
            # 4. Update volume (µm³)
            initial_volume_um3_max = (math.pi/6) * diameter_um_max**3
            # print(f"Initial Volume (max): {initial_volume_um3_max:.2f} µm³")
            new_volume_um3_max = initial_volume_um3_max - volume_loss_um3_max
            # print(f"New Volume (max): {new_volume_um3_max:.2f} µm³")
            
            if new_volume_um3_max <= 0:
                # print(f"Partikel vollständig abradiert nach {step_max} Schritten")
                break
                
            # 5. Calculate new diameter (µm)
            new_diameter_um_max = (6 * new_volume_um3_max / math.pi) ** (1/3)
            
            # Progress output
            """ 
            print(f"Schritt {step_max}:")
            print(f"• Durchmesser (max): {diameter_um_max:.2f} → {new_diameter_um_max:.2f} µm")
            print(f"• Massenverlust (max): {mass_loss_mg_max:.6f} mg/km")
            print(f"• Volumenverlust (max): {volume_loss_um3_max:.2f} µm³")
            print("-"*50)
            """
            # sum up volume loss (max)
            sum_volume_loss_um3_max += volume_loss_um3_max

            # sum up mass loss (max)
            sum_mass_loss_mg_max += mass_loss_mg_max

            # Update for next iteration
            diameter_um_max = new_diameter_um_max # um -> µm

            remaining_km_max -= 1

        # Store results for this polymer
        results[polymer] = {
            'final_diameter_min': diameter_um_min,
            'final_diameter_max': diameter_um_max,
            'total_mass_loss_min': sum_mass_loss_mg_min,
            'total_mass_loss_max': sum_mass_loss_mg_max,
            'total_volume_loss_min': sum_volume_loss_um3_min,
            'total_volume_loss_max': sum_volume_loss_um3_max,
            'remaining_km_min': remaining_km_min,
            'remaining_km_max': remaining_km_max
        }

    return results
    
    """
    print("\nEndergebnis:")
    print(f"MIN: Finaler Durchmesser: {diameter_um_min:.2f} µm")
    print(f"MAX: Finaler Durchmesser: {diameter_um_max:.2f} µm")
    print(f"MIN: Gesamtmassenverlust: {sum_mass_loss_mg_min:.2f} mg")
    print(f"MAX: Gesamtmassenverlust: {sum_mass_loss_mg_max:.2f} mg")
    print(f"MIN: Gesamtvolumenverlust: {sum_volume_loss_um3_min:.2f} µm³")
    print(f"MAX: Gesamtvolumenverlust: {sum_volume_loss_um3_max:.2f} µm³")
    print(f"MIN: Reststrecke: {remaining_km_min:.2f} km")
    print(f"MAX: Reststrecke: {remaining_km_max:.2f} km")
    print("="*50)
    """

def save_results_to_excel(input_file, output_file, all_results):
    """Save results to an Excel file"""
    data = []

    for water_type, polymers_data in all_results.items():
        for polymer, values in polymers_data.items():
            data.append({
                'Water Type': water_type,
                'Polymer': polymer,
                'Final Diameter Min (µm)': values['final_diameter_min'],
                'Final Diameter Max (µm)': values['final_diameter_max'],
                'Total Mass Loss Min (mg)': values['total_mass_loss_min'],
                'Total Mass Loss Max (mg)': values['total_mass_loss_max'],
                'Total Volume Loss Min (µm³)': values['total_volume_loss_min'],
                'Total Volume Loss Max (µm³)': values['total_volume_loss_max'],
                'Remaining Km Min': values['remaining_km_min'],
                'Remaining Km Max': values['remaining_km_max']
            })
    try:
        df = pd.DataFrame(data)
        # Save to Excel
        if input_file == output_file:
            # Append to existing file with new sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                # Check if sheet already exists and delete if necessary
                book = load_workbook(output_file)
                if 'Polymer_Results' in book.sheetnames:
                    del book['Polymer_Results']
                book.save(output_file)
                
                df.to_excel(writer, sheet_name='Polymer_Results', index=False)
        else:
            # Create new file
            df.to_excel(output_file, sheet_name='Polymer_Results', index=False)
    
    except Exception as e:
        print(f"Error saving results to Excel: {e}")
    
def main():
    excel_file = "calculation_results.xlsx"
    data = load_excel_data(excel_file)
    if data is None:
        return
    
    all_results = {}

    counter = 0

    for i, row in data.iterrows():
        water_type = row["Type"]
        total_length = row["Length (km)"]
        w_eff_min = row["Min Power (w/m²)"]
        w_eff_max = row["Max Power (w/m²)"]
        
        if total_length > 0:
            counter += 1

            # print(f"\n*********** {water_type}: {total_length} km *************")
            results = simulate_abrasion(total_length, w_eff_min, w_eff_max)
            all_results[water_type] = results
            # print("="*100)
        else:
            # print(f"Skipping {water_type}")
            continue
    
    # Save results to Excel
    save_results_to_excel(excel_file, "finale_results.xlsx", all_results)
    print("\nResults saved")
    print(f"Simulation completed for {counter} water types.\n")
    
if __name__ == "__main__":
    main()
